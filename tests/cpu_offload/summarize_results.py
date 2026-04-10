#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
# DeepSpeed Team
"""
Collect all results.json files under results/cpu_offload/ and produce a single
summary.xlsx with one sheet per model plus a cross-model summary sheet.

Usage (from DeepSpeed-v0.18.9/):
    python tests/cpu_offload/summarize_results.py

Output:
    results/cpu_offload/summary.xlsx
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ── Locate results root ───────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DS_ROOT = SCRIPT_DIR.parent.parent          # DeepSpeed-v0.18.9/
RESULTS_ROOT = DS_ROOT / "results" / "cpu_offload"
OUTPUT_PATH = RESULTS_ROOT / "summary.xlsx"


def load_all_results():
    """Recursively find results.json files and parse them."""
    records = []
    for path in sorted(RESULTS_ROOT.rglob("results.json")):
        try:
            with open(path) as f:
                data = json.load(f)
            data["_source"] = str(path.relative_to(DS_ROOT))
            records.append(data)
        except Exception as e:
            print(f"[WARN] Could not parse {path}: {e}")
    return records


def model_short(model_str: str) -> str:
    """Return the last path component as a short model name."""
    return model_str.rstrip("/").split("/")[-1]


def style_header(ws, row: int, ncols: int):
    """Apply bold + blue background to a header row."""
    fill = PatternFill(fill_type="solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autofit(ws):
    """Auto-size columns to content."""
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


def write_model_sheet(wb, model_name: str, records: list):
    """One sheet per model: per-step TFLOPS table."""
    ws = wb.create_sheet(title=model_name[:31])  # Excel sheet name limit

    # Header
    headers = ["config", "step", "tflops_per_gpu", "avg_tflops_per_gpu",
               "avg_iter_time_ms", "avg_tokens_per_second", "mode", "mbs",
               "batch_size", "seq_len", "gpus"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for rec in records:
        config = rec.get("config_label") or rec.get("mode", "")
        tflops_list = rec.get("tflops_per_step", [])
        avg_tflops = rec.get("avg_tflops_per_gpu")
        avg_time = rec.get("avg_iter_time_ms")
        avg_tokens = rec.get("avg_tokens_per_second")
        mode = rec.get("mode", "")
        mbs = rec.get("batch_size", "")
        gbs = rec.get("batch_size", "")
        seq = rec.get("seq_len", "")
        gpus = rec.get("gpus", "")

        if tflops_list:
            for step_idx, tflops in enumerate(tflops_list, 1):
                ws.cell(row=row, column=1, value=config)
                ws.cell(row=row, column=2, value=step_idx)
                ws.cell(row=row, column=3, value=tflops)
                ws.cell(row=row, column=4, value=avg_tflops)
                ws.cell(row=row, column=5, value=avg_time)
                ws.cell(row=row, column=6, value=avg_tokens)
                ws.cell(row=row, column=7, value=mode)
                ws.cell(row=row, column=8, value=mbs)
                ws.cell(row=row, column=9, value=gbs)
                ws.cell(row=row, column=10, value=seq)
                ws.cell(row=row, column=11, value=gpus)
                row += 1
        else:
            # MoE or missing per-step data: write summary row with step=N/A
            ws.cell(row=row, column=1, value=config)
            ws.cell(row=row, column=2, value="N/A")
            ws.cell(row=row, column=3, value=None)
            ws.cell(row=row, column=4, value=avg_tflops)
            ws.cell(row=row, column=5, value=avg_time)
            ws.cell(row=row, column=6, value=avg_tokens)
            ws.cell(row=row, column=7, value=mode)
            ws.cell(row=row, column=8, value=mbs)
            ws.cell(row=row, column=9, value=gbs)
            ws.cell(row=row, column=10, value=seq)
            ws.cell(row=row, column=11, value=gpus)
            row += 1

    autofit(ws)


def write_summary_sheet(wb, all_records: list):
    """Cross-model summary: one row per config, showing avg metrics."""
    ws = wb.create_sheet(title="Summary", index=0)

    headers = ["model", "config", "mode", "mbs", "avg_tflops_per_gpu",
               "avg_iter_time_ms", "avg_tokens_per_second",
               "pipeline_bubble_fraction", "gpus", "seq_len", "source"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    for row, rec in enumerate(all_records, 2):
        ws.cell(row=row, column=1, value=model_short(rec.get("model", "")))
        ws.cell(row=row, column=2, value=rec.get("config_label") or rec.get("mode", ""))
        ws.cell(row=row, column=3, value=rec.get("mode", ""))
        ws.cell(row=row, column=4, value=rec.get("batch_size"))
        ws.cell(row=row, column=5, value=rec.get("avg_tflops_per_gpu"))
        ws.cell(row=row, column=6, value=rec.get("avg_iter_time_ms"))
        ws.cell(row=row, column=7, value=rec.get("avg_tokens_per_second"))
        ws.cell(row=row, column=8, value=rec.get("pipeline_bubble_fraction"))
        ws.cell(row=row, column=9, value=rec.get("gpus"))
        ws.cell(row=row, column=10, value=rec.get("seq_len"))
        ws.cell(row=row, column=11, value=rec.get("_source", ""))

    autofit(ws)


def main():
    records = load_all_results()
    if not records:
        print(f"No results.json found under {RESULTS_ROOT}")
        return

    print(f"Found {len(records)} result(s).")

    # Group by model
    by_model: dict[str, list] = {}
    for rec in records:
        key = model_short(rec.get("model", "unknown"))
        by_model.setdefault(key, []).append(rec)

    wb = openpyxl.Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    write_summary_sheet(wb, records)

    for model_name, recs in sorted(by_model.items()):
        write_model_sheet(wb, model_name, recs)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
